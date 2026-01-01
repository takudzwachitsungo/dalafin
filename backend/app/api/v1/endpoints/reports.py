from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func, extract, and_
from datetime import datetime, timedelta, date
from typing import Literal
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from decimal import Decimal

from app.core.database import get_db
from app.models.user import User
from app.models.transaction import Transaction
from app.models.income import Income
from app.api.v1.endpoints.auth import get_current_user

router = APIRouter()


def get_date_range(period: str):
    """Calculate start and end dates based on period"""
    today = datetime.now()
    
    if period == "weekly":
        start_date = today - timedelta(days=7)
    elif period == "monthly":
        start_date = today.replace(day=1)
    elif period == "quarterly":
        current_quarter = (today.month - 1) // 3
        start_month = current_quarter * 3 + 1
        start_date = today.replace(month=start_month, day=1)
    elif period == "yearly":
        start_date = today.replace(month=1, day=1)
    else:
        start_date = today - timedelta(days=30)
    
    return start_date, today


def create_styled_excel(period: str, transactions, incomes, user_email: str):
    """Create a beautifully styled Excel report"""
    wb = Workbook()
    wb.remove(wb.active)
    
    # Define styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(name="Arial", size=12, bold=True, color="FFFFFF")
    title_font = Font(name="Arial", size=16, bold=True, color="2F5496")
    subtotal_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    subtotal_font = Font(name="Arial", size=11, bold=True)
    border = Border(
        left=Side(style='thin', color='B4C7E7'),
        right=Side(style='thin', color='B4C7E7'),
        top=Side(style='thin', color='B4C7E7'),
        bottom=Side(style='thin', color='B4C7E7')
    )
    
    # Summary Sheet
    ws_summary = wb.create_sheet("Summary")
    ws_summary.merge_cells('A1:D1')
    ws_summary['A1'] = f"Financial Report - {period.title()}"
    ws_summary['A1'].font = title_font
    ws_summary['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws_summary.row_dimensions[1].height = 30
    
    ws_summary['A3'] = "Report Generated:"
    ws_summary['B3'] = datetime.now().strftime("%B %d, %Y at %I:%M %p")
    ws_summary['A4'] = "User:"
    ws_summary['B4'] = user_email
    
    # Calculate totals
    total_expenses = sum(float(t.amount) for t in transactions)
    total_income = sum(float(i.amount) for i in incomes)
    net_balance = total_income - total_expenses
    
    ws_summary['A6'] = "Financial Summary"
    ws_summary['A6'].font = Font(name="Arial", size=12, bold=True)
    
    ws_summary['A8'] = "Total Income"
    ws_summary['B8'] = total_income
    ws_summary['B8'].number_format = '$#,##0.00'
    ws_summary['B8'].font = Font(color="008000", bold=True)
    
    ws_summary['A9'] = "Total Expenses"
    ws_summary['B9'] = total_expenses
    ws_summary['B9'].number_format = '$#,##0.00'
    ws_summary['B9'].font = Font(color="FF0000", bold=True)
    
    ws_summary['A10'] = "Net Balance"
    ws_summary['B10'] = net_balance
    ws_summary['B10'].number_format = '$#,##0.00'
    ws_summary['B10'].font = Font(color="0000FF" if net_balance >= 0 else "FF0000", bold=True, size=12)
    
    # Category breakdown
    category_totals = {}
    for t in transactions:
        category_totals[t.category] = category_totals.get(t.category, 0) + float(t.amount)
    
    ws_summary['A12'] = "Expenses by Category"
    ws_summary['A12'].font = Font(name="Arial", size=12, bold=True)
    
    row = 14
    for category, amount in sorted(category_totals.items(), key=lambda x: x[1], reverse=True):
        ws_summary[f'A{row}'] = category.title()
        ws_summary[f'B{row}'] = amount
        ws_summary[f'B{row}'].number_format = '$#,##0.00'
        ws_summary[f'C{row}'] = amount / total_expenses if total_expenses > 0 else 0
        ws_summary[f'C{row}'].number_format = '0.0%'
        row += 1
    
    # Set column widths
    ws_summary.column_dimensions['A'].width = 25
    ws_summary.column_dimensions['B'].width = 20
    ws_summary.column_dimensions['C'].width = 15
    
    # Transactions Sheet
    ws_transactions = wb.create_sheet("Transactions")
    ws_transactions.merge_cells('A1:F1')
    ws_transactions['A1'] = "Expense Transactions"
    ws_transactions['A1'].font = title_font
    ws_transactions['A1'].alignment = Alignment(horizontal='center')
    ws_transactions.row_dimensions[1].height = 25
    
    # Headers
    headers = ["Date", "Category", "Amount", "Note", "Impulse", "Time"]
    for col, header in enumerate(headers, 1):
        cell = ws_transactions.cell(row=3, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Data
    row = 4
    for t in sorted(transactions, key=lambda x: x.date, reverse=True):
        ws_transactions[f'A{row}'] = t.date.strftime("%Y-%m-%d")
        ws_transactions[f'B{row}'] = t.category.title()
        ws_transactions[f'C{row}'] = float(t.amount)
        ws_transactions[f'C{row}'].number_format = '$#,##0.00'
        ws_transactions[f'D{row}'] = t.note or ""
        ws_transactions[f'E{row}'] = "Yes" if t.is_impulse else "No"
        ws_transactions[f'F{row}'] = t.date.strftime("%I:%M %p")
        
        for col in range(1, 7):
            ws_transactions.cell(row=row, column=col).border = border
        
        row += 1
    
    # Total row
    ws_transactions[f'A{row}'] = "TOTAL"
    ws_transactions[f'A{row}'].font = subtotal_font
    ws_transactions[f'A{row}'].fill = subtotal_fill
    ws_transactions[f'C{row}'] = total_expenses
    ws_transactions[f'C{row}'].number_format = '$#,##0.00'
    ws_transactions[f'C{row}'].font = subtotal_font
    ws_transactions[f'C{row}'].fill = subtotal_fill
    
    # Set column widths
    ws_transactions.column_dimensions['A'].width = 15
    ws_transactions.column_dimensions['B'].width = 15
    ws_transactions.column_dimensions['C'].width = 15
    ws_transactions.column_dimensions['D'].width = 30
    ws_transactions.column_dimensions['E'].width = 10
    ws_transactions.column_dimensions['F'].width = 12
    
    # Income Sheet
    ws_income = wb.create_sheet("Income")
    ws_income.merge_cells('A1:E1')
    ws_income['A1'] = "Income Records"
    ws_income['A1'].font = title_font
    ws_income['A1'].alignment = Alignment(horizontal='center')
    ws_income.row_dimensions[1].height = 25
    
    # Headers
    income_headers = ["Date", "Source", "Amount", "Description", "Time"]
    for col, header in enumerate(income_headers, 1):
        cell = ws_income.cell(row=3, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Data
    row = 4
    for i in sorted(incomes, key=lambda x: x.date, reverse=True):
        ws_income[f'A{row}'] = i.date.strftime("%Y-%m-%d")
        ws_income[f'B{row}'] = i.source
        ws_income[f'C{row}'] = float(i.amount)
        ws_income[f'C{row}'].number_format = '$#,##0.00'
        ws_income[f'D{row}'] = i.description or ""
        ws_income[f'E{row}'] = i.date.strftime("%I:%M %p")
        
        for col in range(1, 6):
            ws_income.cell(row=row, column=col).border = border
        
        row += 1
    
    # Total row
    ws_income[f'A{row}'] = "TOTAL"
    ws_income[f'A{row}'].font = subtotal_font
    ws_income[f'A{row}'].fill = subtotal_fill
    ws_income[f'C{row}'] = total_income
    ws_income[f'C{row}'].number_format = '$#,##0.00'
    ws_income[f'C{row}'].font = subtotal_font
    ws_income[f'C{row}'].fill = subtotal_fill
    
    # Set column widths
    ws_income.column_dimensions['A'].width = 15
    ws_income.column_dimensions['B'].width = 20
    ws_income.column_dimensions['C'].width = 15
    ws_income.column_dimensions['D'].width = 35
    ws_income.column_dimensions['E'].width = 12
    
    return wb


@router.get("/export")
async def export_report(
    period: Literal["weekly", "monthly", "quarterly", "yearly"] = Query("monthly"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Generate and download Excel report for specified period"""
    start_date, end_date = get_date_range(period)
    
    # Fetch transactions
    transactions = db.query(Transaction).filter(
        and_(
            Transaction.user_id == current_user.id,
            Transaction.date >= start_date,
            Transaction.date <= end_date
        )
    ).all()
    
    # Fetch incomes
    incomes = db.query(Income).filter(
        and_(
            Income.user_id == current_user.id,
            Income.date >= start_date,
            Income.date <= end_date
        )
    ).all()
    
    # Create Excel file
    wb = create_styled_excel(period, transactions, incomes, current_user.email)
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Generate filename
    filename = f"Financial_Report_{period.title()}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
